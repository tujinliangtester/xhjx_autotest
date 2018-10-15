from openpyxl import Workbook

from openpyxl import load_workbook

from openpyxl.writer.excel import ExcelWriter
import HtmlTestRunner
import unittest
class bd(unittest.TestCase):
    def do(self):
        wb=load_workbook('E:\\tjl个人\\study\\auto\\笔记.xlsx')
        ws=wb.worksheets[0]
        str=ws.cell(row=3,column=1).value
        print(str)
        ws2=wb.worksheets[1]
        ws2['A1']='aaaa什么鬼'
        ws2['B1']=str
        wb.save('E:\\tjl个人\\study\\auto\\笔记2.xlsx')
        print('ok')
    def do2(self):
        print('11')
if __name__=="__main__":
    testunit=unittest.TestSuite()
    testunit.addTest(bd('do'))
    testunit.addTest(bd('do2'))
    fp=open("result"+".html",'w')
    runner=HtmlTestRunner.HTMLTestRunner(output='E:\\tjl个人\\study\\auto\\',stream=fp,report_title='test result',descriptions=u'result:')
    runner.run(testunit)
    fp.close()