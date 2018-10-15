import xlrd
import xlwt
import re
 # import openpyxl
# from openpyxl import load_workbook
from openpyxl import Workbook
#写入excel使用(支持07)
from openpyxl import load_workbook
# data=xlrd.open_workbook('E:\\tjl个人\\测试过程\\微商城3.0\\订单\\1.xlsx')
# table=data.sheet_by_name('1')
data=xlwt.Workbook
table=data.add_sheet('1')
# if not table:
#     print('not found')

file=open('E:\\tjl个人\\测试过程\\微商城3.0\\订单\\草稿.txt')
i=1
while(1):
    line=file.readline()
    print(line)
    i=i+1
    if not line:
        break
    # table.cell(i,1).value=line
    # table.put_cell(i,1,1,line,0)

    # table.cell(i,1).value=line
    print(table.cell(i, 1))

    # table.put_cell_ragged(i,1,1,line,0)





